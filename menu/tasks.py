

from menu.parser import parse_data, fetch_yml_data
# menu/tasks.py
from celery import shared_task, chain # Importer chain
from celery.exceptions import MaxRetriesExceededError, Ignore
from django.utils import timezone
from django.db import transaction
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.conf import settings
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
# from django.urls import reverse
from collections import defaultdict
import csv
import io
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter # Assurez-vous que c'est importé
import logging
import os
# Importer vos modèles et fonctions
from .models import FeedLink, Product, MenuCatalog, AvailabilityReport # Importer tous les modèles
# Adaptez ces imports selon l'emplacement de vos fonctions fetch/parse
# from .parser import fetch_yml_data, parse_data, download_and_save_image, generate_unique_slug

logger = logging.getLogger(__name__)

# --- Constantes ---
UNCATEGORIZED_CATEGORY_ID = getattr(settings, 'UNCATEGORIZED_CATEGORY_ID', '2')
MARKETING_ANALYST_EMAIL = getattr(settings, 'MARKETING_ANALYST_EMAIL', None)
# Fallback si aucun rapport n'existe (en heures)
DEFAULT_REPORT_INITIAL_LOOKBACK_HOURS = getattr(settings, 'DEFAULT_REPORT_INITIAL_LOOKBACK_HOURS', 1)
# Fallback si erreur de récupération du dernier rapport
DEFAULT_FALLBACK_LOOKBACK_MINUTES = 12
# Dossier pour les rapports
REPORT_UPLOAD_SUBDIR = getattr(settings, 'REPORT_UPLOAD_DIR', 'uploads/reports/availability')


# --- Assurez-vous que fetch_yml_data et parse_data sont définis/importés ---
# def fetch_yml_data(feed_url): ...
# def parse_data(data=None, yml_feed_url=None): ... # Doit sauvegarder les produits et mettre à jour le timestamp

# ---------------------------------------------------------------------------
# Tâche 1: Mise à Jour du Flux (Retourne indicateur de parsing)
# ---------------------------------------------------------------------------
@shared_task(bind=True, ignore_result=True, max_retries=3, default_retry_delay=60) # ignore_result=False
def update_feed_data(self, feed_url):
    """
    Met à jour les données depuis un flux YML.
    Retourne True si de nouvelles données ont été traitées (parsing tenté), False sinon.
    """
    logger.info(f"[{feed_url}] Starting update task (part of chain).")
    parsing_attempted = False
    update_successful = False
    feed_link_obj = None

    try:
        feed_link_obj = FeedLink.objects.get(feedlink=feed_url)
        data = fetch_yml_data(feed_url)

        if data:
            parsing_attempted = True # MARQUEUR IMPORTANT
            logger.info(f"[{feed_url}] New data fetched. Parsing and saving...")
            with transaction.atomic():
                parsed_result = parse_data(data=data, yml_feed_url=feed_url)

            if parsed_result is not None: # Ou une condition plus spécifique si besoin
                logger.info(f"[{feed_url}] Parsing and saving process completed.")
                update_successful = True
            else:
                logger.info(f"[{feed_url}] Parsing function completed but returned None or empty.")
                update_successful = True # Processus terminé sans erreur critique
        else:
            logger.info(f"[{feed_url}] No new data fetched (ETag unchanged or fetch error).")
            update_successful = True # Vérification ETag OK

    except FeedLink.DoesNotExist as e:
        logger.error(f"[{feed_url}] FeedLink object does not exist. Task failed definitively.")
        raise e # Fait échouer explicitement la tâche et la chaîne

    except Exception as e:
        logger.exception(f"[{feed_url}] Error during feed update task: {e}")
        try:
            raise self.retry(exc=e, countdown=int(self.default_retry_delay * (2 ** self.request.retries)))
        except MaxRetriesExceededError:
            logger.error(f"[{feed_url}] Max retries exceeded. Task failed definitively.")
            raise # Fait échouer explicitement la tâche et la chaîne
        update_successful = False

    finally:
        if feed_link_obj:
            try:
                feed_link_obj.updated_at = timezone.now()
                feed_link_obj.last_run_status = 'success' if update_successful else 'failure'
                feed_link_obj.save(update_fields=['updated_at', 'last_run_status'])
            except Exception as e_save:
                logger.error(f"[{feed_url}] Error updating FeedLink final status: {e_save}")

    logger.info(f"[{feed_url}] Task finished. Parsing attempted: {parsing_attempted}")
    return parsing_attempted # Retourne True si le parsing a été tenté


# ---------------------------------------------------------------------------
# Tâche 2: Notification des Changements (Basée sur dernier rapport réussi)
# ---------------------------------------------------------------------------
@shared_task(bind=True, ignore_result=True, max_retries=2, default_retry_delay=300)
def notify_availability_changes(self, parsing_attempted_in_previous_task, recipient_email=None):
    """
    Génère un rapport Excel (.xlsx) des changements depuis le DERNIER RAPPORT RÉUSSI
    et envoie un email avec un lien.
    Ne s'exécute que si la tâche précédente indique qu'un parsing a eu lieu.
    """
    # --- Vérification Initiale Basée sur la Tâche Précédente ---
    if not parsing_attempted_in_previous_task:
        logger.info("Skipping notification because previous task indicated no new data was parsed.")
        return "Skipped: No new data parsed."
    # ----------------------------------------------------------

    logger.info("Availability change notification task started (triggered by chain).")

    # Détermination Destinataire
    if not recipient_email:
        recipient_email = MARKETING_ANALYST_EMAIL
        if not recipient_email:
            logger.error("Recipient email not configured. Task cannot proceed.")
            return "Configuration Error: Recipient email not set."

    # --- 1. Trouver l'heure de début (dernier rapport NOTIFIED) ---
    last_success_time = None
    try:
        last_successful_report = AvailabilityReport.objects.filter(
            status=AvailabilityReport.STATUS_NOTIFIED
        ).order_by('-generated_at').only('generated_at').first()
        if last_successful_report:
            last_success_time = last_successful_report.generated_at
        else:
            last_success_time = timezone.now() - timezone.timedelta(hours=DEFAULT_REPORT_INITIAL_LOOKBACK_HOURS)
    except Exception as e_last_report:
        logger.error(f"Error fetching last report time: {e_last_report}. Falling back.")
        last_success_time = timezone.now() - timezone.timedelta(minutes=DEFAULT_FALLBACK_LOOKBACK_MINUTES)

    start_time = last_success_time
    start_time_iso = start_time.isoformat()
    logger.info(f"Fetching changes since last successful report: {start_time_iso}")

    # --- 2. Récupération Produits Modifiés ---
    try:
        changed_products_qs = Product.objects.filter(
            availability_changed_at__gte=start_time,
        ).exclude(
            catalog__id=UNCATEGORIZED_CATEGORY_ID
        ).select_related('catalog').only(
            'id', 'title', 'slug', 'available', 'catalog__id', 'catalog__name'
        ).order_by('catalog__name', 'title')

        change_count = changed_products_qs.count()
        if change_count == 0:
            logger.info(f"No new classified product availability changes detected since {start_time_iso}.")
            # Optionnel: Créer un rapport vide ?
            # AvailabilityReport.objects.create(status=AvailabilityReport.STATUS_NOTIFIED, product_change_count=0)
            return "No new changes since last successful report."
        changed_product_ids = list(changed_products_qs.values_list('id', flat=True))

    except Exception as e_db: # ... (Gestion erreur DB + retry) ...
        logger.exception(f"Database error fetching changed products: {e_db}")
        try: raise self.retry(exc=e_db, countdown=60)
        except MaxRetriesExceededError: return f"Failed: DB error - {e_db}"
        except Exception as retry_e: return f"Failed: Error retrying DB fetch - {retry_e}"

    # --- 3. Créer l'enregistrement initial du Rapport ---
    report_obj = None
    try:
        report_obj = AvailabilityReport.objects.create(
            status=AvailabilityReport.STATUS_PENDING,
            product_change_count=change_count
        )
    except Exception as e_report_init:
        logger.error(f"CRITICAL: Failed to create initial AvailabilityReport: {e_report_init}", exc_info=True)
        return f"Failed: Could not create report object - {e_report_init}"

    # --- 4. Génération et Sauvegarde du Rapport Excel ---
    # ... (Logique de génération/sauvegarde Excel identique à la réponse précédente) ...
    # ... (Met à jour report_obj avec infos fichier + status GENERATED/FAILED) ...
    report_url = None; report_filename = None; saved_file_path = None
    newly_unavailable_count = 0; newly_available_count = 0
    generation_successful = False
    try:
        excel_buffer = io.BytesIO(); workbook = openpyxl.Workbook(); sheet = workbook.active
        sheet.title = "Изменения"; headers = ['Категория', 'Статус', 'ID', 'Название', 'Slug']
        sheet.append(headers); header_font = Font(bold=True); center_alignment = Alignment(horizontal='center')
        for col,w in enumerate([35,18,15,60,50],1): sheet.column_dimensions[get_column_letter(col)].width=w
        for col,h in enumerate(headers,1): c=sheet.cell(row=1,column=col); c.font=header_font; c.alignment=center_alignment
        for product in changed_products_qs.iterator(chunk_size=2000):
            cat = product.catalog.name if product.catalog else "Без категории"
            stat = "В наличии" if product.available else "Нет в наличии"
            if product.available: newly_available_count += 1
            else: newly_unavailable_count += 1
            sheet.append([cat, stat, product.id, product.title, product.slug])
        workbook.save(excel_buffer); excel_buffer.seek(0)
        ts = timezone.now().strftime('%Y%m%d_%H%M%S'); fn = f"izmeneniya_dostupnosti_{ts}.xlsx"
        fpath = os.path.join(REPORT_UPLOAD_SUBDIR, fn)
        saved_file_path = default_storage.save(fpath, ContentFile(excel_buffer.read()))
        report_filename = fn; excel_buffer.close(); logger.info(f"Report generated: {saved_file_path}")
        generation_successful = True
        report_obj.file_path=saved_file_path; report_obj.filename=report_filename
        report_obj.unavailable_count=newly_unavailable_count; report_obj.available_count=newly_available_count
        report_obj.status=AvailabilityReport.STATUS_GENERATED; report_obj.generated_at=timezone.now()
        report_obj.error_message=None; report_obj.save()
        logger.info(f"Report {report_obj.id} updated to GENERATED.")
        try: report_url = default_storage.url(saved_file_path)
        except Exception: report_url = None
    except Exception as e_file:
        logger.exception(f"CRITICAL: Error generating/saving report {report_obj.id}: {e_file}")
        report_obj.status = AvailabilityReport.STATUS_FAILED; report_obj.error_message = str(e_file)[:1000]
        report_obj.generated_at = timezone.now(); report_obj.save()

    # --- 5. Envoi de l'Email (si génération OK) ---
    email_sent_successfully = False
    if generation_successful:
        # ... (Logique d'envoi email identique à la réponse précédente) ...
        # ... (Met à jour report_obj.status à NOTIFIED ou FAILED) ...
        subject = f"Изменения доступности ({timezone.now():%Y-%m-%d %H:%M})"
        context = {'newly_unavailable_count': newly_unavailable_count, 'newly_available_count': newly_available_count,
                   'total_changes': change_count, 'report_url': report_url, 'report_filename': report_filename,
                   'start_time_iso': start_time_iso, 'update_time': timezone.now(), 'report_admin_url': None}
        html_message = render_to_string('checkout/elements/emails/availability_change_notification.html', context)
        try:
            send_mail(subject, '', settings.DEFAULT_FROM_EMAIL, [recipient_email], html_message=html_message)
            logger.info(f"Report link email sent for report {report_obj.id}.")
            report_obj.status = AvailabilityReport.STATUS_NOTIFIED
            report_obj.save(update_fields=['status'])
            email_sent_successfully = True
        except Exception as e_mail:
            logger.exception(f"Failed send report email {report_obj.id}: {e_mail}")
            report_obj.status = AvailabilityReport.STATUS_FAILED
            report_obj.error_message = f"Email Error: {str(e_mail)[:500]}"
            report_obj.save(update_fields=['status', 'error_message'])
            try: raise self.retry(exc=e_mail, countdown=300)
            except MaxRetriesExceededError: logger.error(f"Max retries email report {report_obj.id}.")
            except Exception as retry_e: logger.error(f"Error retrying email {report_obj.id}: {retry_e}")

    # --- 6. Associer les produits (Optionnel) ---
    if report_obj and report_obj.status in [AvailabilityReport.STATUS_GENERATED, AvailabilityReport.STATUS_NOTIFIED]:
        try: report_obj.notified_products.set(changed_product_ids)
        except Exception as e_m2m: logger.error(f"Failed associate products report {report_obj.id}: {e_m2m}")

    # --- Retour final ---
    if report_obj.status == AvailabilityReport.STATUS_NOTIFIED: return f"Success: Report {report_obj.id} sent."
    elif report_obj.status == AvailabilityReport.STATUS_GENERATED: return f"Partial Success: Report {report_obj.id} generated, email failed."
    else: return f"Failed: Report {report_obj.id} failed (Status: {report_obj.status})."


# --- Tâche Périodique Principale (Utilise la chaîne simple) ---
@shared_task(name='menu.tasks.update_all_feeds')
def update_all_feeds():
    """Lance une CHAÎNE update->notify pour chaque flux activé."""
    logger.info("Starting periodic scheduling of simple update chains...")
    active_feed_links = FeedLink.objects.filter(enabled=True)
    count = 0
    for feed_link in active_feed_links:
        logger.info(f"Scheduling simple update chain for: {feed_link.feedlink} (ID: {feed_link.id})")
        try:
            task_chain = chain(
                update_feed_data.s(feed_link.feedlink),
                # La tâche suivante reçoit le retour (parsing_attempted)
                # et n'a plus besoin de lookback_minutes car elle utilise le dernier rapport
                notify_availability_changes.s()
            )
            task_chain.delay()
            count += 1
        except Exception as e_chain:
            logger.error(f"Failed to schedule simple chain for feed {feed_link.feedlink}: {e_chain}", exc_info=True)
    logger.info(f"Scheduled {count} simple update chains.")