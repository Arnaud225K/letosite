import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import io
from datetime import datetime
from collections import defaultdict
import logging
import re

from django.shortcuts import render, redirect, get_object_or_404
from django.views import View
from django.http import HttpResponse, JsonResponse, Http404
from django.contrib.admin.views.decorators import staff_member_required
from django.utils.decorators import method_decorator
from django.urls import reverse
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django import forms
from django.contrib import messages
from django.db import transaction, IntegrityError
from django.db.models import Q, Prefetch, Count
from django.utils.text import slugify
from unidecode import unidecode
from django.contrib import admin
from django.utils import timezone
from django.conf import settings

# Import Select2 widgets si disponible
try:
    from django_select2.forms import Select2Widget, Select2MultipleWidget
    DJANGO_SELECT2_INSTALLED = True
except ImportError:
    Select2Widget = forms.Select
    Select2MultipleWidget = forms.SelectMultiple # Fallback
    DJANGO_SELECT2_INSTALLED = False

# Importer les modèles de l'application 'menu'
from menu.models import Product, MenuCatalog, FilterCategory, FilterValue, TypeMenu


UNCATEGORIZED_CATEGORY_ID = getattr(settings, 'UNCATEGORIZED_CATEGORY_ID', '2')
PRODUCT_CATEGORY_TYPE_IDS = getattr(settings, 'VALID_ASSIGN_CATEGORY_TYPE_IDS', ['7', '8'])
REPORT_UPLOAD_SUBDIR = 'uploads/reports/availability'


logger = logging.getLogger(__name__)


# --- Formulaires utilisés par ces vues ---
class AdminExportByCategoryForm(forms.Form):
    category = forms.ModelChoiceField(
        queryset=MenuCatalog.objects.filter(is_hidden=False, type_menu_id__in=PRODUCT_CATEGORY_TYPE_IDS).order_by('name'),
        
        label="Категория продуктов для экспорта",
        required=True,
        widget=Select2Widget(attrs={'data-placeholder': 'Выберите категорию...', 'class': 'form-control'}),
        empty_label=None
    )
    filter_categories_to_export = forms.ModelMultipleChoiceField(
        queryset=FilterCategory.objects.filter(is_active=True).order_by('order', 'name'),
        label="Включить/Добавить колонки фильтров",
        required=False,
        widget=Select2MultipleWidget(attrs={'data-placeholder': 'Авто (используемые) + ваши...', 'class': 'form-control'}),
        help_text="Выберите фильтры для включения. Используемые фильтры будут добавлены автоматически, если ничего не выбрано."
    )


class AdminImportFileForm(forms.Form):
    file = forms.FileField(
        label="Файл Excel (.xlsx)",
        required=True,
        widget=forms.FileInput(attrs={'accept': '.xlsx,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'})
    )



# --- Fonctions Utilitaires (Optionnel - pour get_base_admin_context) ---
def get_base_admin_context(request, target_model=Product):
    """Helper pour obtenir le contexte admin de base."""
    # Tente d'obtenir le contexte standard si possible
    context = {}
    if hasattr(admin.site, 'each_context'):
        context = admin.site.each_context(request)

    # Assure que les éléments clés sont présents
    context.update({
        'has_permission': request.user.is_staff, # Vérification basique
        'opts': target_model._meta,
        'app_label': target_model._meta.app_label,
        'site_header': getattr(admin.site, 'site_header', 'Django administration'),
        'site_title': getattr(admin.site, 'site_title', 'Django site admin'),
        'user': request.user,
        'is_nav_sidebar_enabled': True, # Suppose la sidebar admin
    })
    return context



# --- Vues ---
@method_decorator(staff_member_required, name='dispatch')
class ProductExportSetupView(View):
    """Affiche le formulaire pour choisir la catégorie et les filtres."""
    form_class = AdminExportByCategoryForm
    template_name = 'admin_m/product_export_setup.html'

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        context = { 
                **get_base_admin_context(request, Product), 
                'title': 'Настройка Экспорта Продуктов',
                'form': form, 'media': form.media 
            }
        return render(request, self.template_name, context)

# --- Vues Personnalisées ---
@method_decorator(staff_member_required, name='dispatch')
class ProductExportDownloadView(View):
    """Génère et télécharge le fichier Excel basé sur les choix."""

    def get(self, request, *args, **kwargs):
        form = AdminExportByCategoryForm(request.GET)
        if not form.is_valid():
            logger.warning(f"Invalid export GET parameters: {form.errors.as_json()}")
            messages.error(request, "Неверные параметры для экспорта.")
            return redirect('admin_m:product_export_setup')

        category = form.cleaned_data['category']
        user_selected_extra_filter_cats_qs = form.cleaned_data['filter_categories_to_export']

        logger.info(f"Generating Excel export for category: '{category.name}'.")

        # 1. Récupère les produits
        products_to_export = Product.objects.filter(catalog=category).order_by('sku')

        if not products_to_export.exists():
            messages.warning(request, f"В категории '{category.name}' нет продуктов для экспорта.")
            return redirect('admin_m:product_export_setup')

        # --- 2. Déterminer les FilterCategory à inclure (CORRECTION) ---
        # Utilise la méthode des IDs en Python pour éviter l'erreur de combinaison de QuerySet

        # 2a. IDs des catégories utilisées par les produits exportés
        used_filter_cat_ids = set(
            FilterCategory.objects.filter(
                is_active=True,
                values__products__in=products_to_export
            ).values_list('pk', flat=True) # Récupère seulement les PKs
        )
        logger.debug(f"Used filter category IDs in category {category.pk}: {used_filter_cat_ids}")

        # 2b. IDs des catégories explicitement sélectionnées par l'utilisateur
        user_selected_ids = set(user_selected_extra_filter_cats_qs.values_list('pk', flat=True))
        logger.debug(f"User selected extra filter category IDs: {user_selected_ids}")

        # 2c. Union des deux ensembles d'IDs
        final_filter_cat_ids = used_filter_cat_ids.union(user_selected_ids)
        logger.debug(f"Final unique filter category IDs for export: {final_filter_cat_ids}")

        # 3. Récupère les objets FilterCategory finaux, triés
        filter_categories_for_export = FilterCategory.objects.filter(
            pk__in=final_filter_cat_ids
        ).order_by('order', 'name')
        logger.info(f"Final filter categories to include in export: {[fc.name for fc in filter_categories_for_export]}")
        # --- Fin Détermination Filtres (Corrigée) ---


        # --- 4. Précharger les filtres pertinents ---
        if filter_categories_for_export.exists():
            products_to_export = products_to_export.prefetch_related(
                Prefetch(
                    'filters',
                    queryset=FilterValue.objects.filter(category__in=filter_categories_for_export).select_related('category'),
                    to_attr='prefetched_filters_for_export'
                )
            )
        else:
            products_to_export = products_to_export.only('sku', 'title', 'catalog_id', 'ed_izm')

        # --- 5. Génération Excel ---
        try:
            excel_buffer = io.BytesIO()
            workbook = openpyxl.Workbook(); sheet = workbook.active
            sheet.title = slugify(unidecode(category.name))[:30]

            # En-têtes: Base + Filtres Finaux
            headers = ['Артикул (SKU)', 'Название', 'ID Категории (для изменения)', 'Ед. изм.']
            filter_cat_slug_to_header = {}
            for fc in filter_categories_for_export: # Utilise la liste finale
                column_name = f'Фильтр: {fc.name} ({fc.slug})'
                headers.append(column_name)
                filter_cat_slug_to_header[fc.slug] = column_name
            sheet.append(headers)
            # Styles Headers (Exemple)
            header_font=Font(bold=True); center_alignment=Alignment(horizontal='center')
            col_widths = {'A': 25, 'B': 45, 'C': 15, 'D': 12}
            for col_num, header_title in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                sheet.column_dimensions[col_letter].width = col_widths.get(col_letter, 30)
                cell=sheet.cell(row=1,column=col_num); cell.font=header_font; cell.alignment=center_alignment

            # Données Produit
            for product in products_to_export.iterator(chunk_size=1000):
                product_filters_by_slug = defaultdict(list)
                filters_to_iterate = getattr(product, 'prefetched_filters_for_export', [])
                for fv in filters_to_iterate:
                    if fv.category.slug in filter_cat_slug_to_header:
                        product_filters_by_slug[fv.category.slug].append(fv.value)
                row_data = [ product.sku, product.title, product.catalog_id, product.ed_izm ]
                for fc in filter_categories_for_export: # Maintient l'ordre
                    values = product_filters_by_slug.get(fc.slug, [])
                    row_data.append("|".join(sorted(values)))
                sheet.append(row_data)

            workbook.save(excel_buffer); excel_buffer.seek(0)

            # --- 6. Création Réponse HTTP ---
            response = HttpResponse(
                excel_buffer.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            timestamp = datetime.now().strftime('%Y%m%d_%H%M')
            filename = f"EXPORT_cat_{slugify(unidecode(category.name))}_{timestamp}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            logger.info(f"Excel export generated successfully: {filename}")
            excel_buffer.close()
            return response

        except Exception as e_excel:
            logger.exception(f"Error generating Excel export for category {category.name}: {e_excel}")
            messages.error(request, "Произошла ошибка при создании файла Excel.")
            return redirect('admin_m:product_export_setup')

# --- Nouvelle Vue AJAX pour pré-sélection ---
@staff_member_required
def ajax_get_relevant_filters(request, category_id):
    """Retourne les IDs des FilterCategory utilisées par les produits."""
    if not request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'error': 'AJAX only'}, status=400)
    try:
        # Pas besoin de l'objet catégorie complet ici
        product_pks = Product.objects.filter(catalog_id=category_id).values_list('pk', flat=True)
        if not product_pks.exists():
            return JsonResponse({'relevant_filter_ids': []})

        # Trouve les IDs de FilterCategory actives utilisées
        relevant_filter_cat_ids = list(
            FilterCategory.objects.filter(
                is_active=True,
                values__products__pk__in=list(product_pks) # Recherche rapide par PK produit
            ).distinct().values_list('pk', flat=True)
        )
        logger.debug(f"AJAX: Relevant filter category IDs for category {category_id}: {relevant_filter_cat_ids}")
        return JsonResponse({'relevant_filter_ids': relevant_filter_cat_ids})

    except Exception as e:
        logger.error(f"Error fetching relevant filters AJAX for category {category_id}: {e}", exc_info=True)
        return JsonResponse({'error': 'Server error'}, status=500)


@method_decorator(staff_member_required, name='dispatch')
class ProductImportView(View):
    """Affiche le formulaire d'upload."""
    form_class = AdminImportFileForm
    template_name = 'admin_m/product_import_setup.html'

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        context = {
            **get_base_admin_context(request, Product),
            'title': 'Импорт (Категории/Фильтры)',
            'form': form,
            'media': form.media,
        }
        return render(request, self.template_name, context)




@method_decorator(staff_member_required, name='dispatch')
@method_decorator(transaction.atomic, name='dispatch')
class ProductImportProcessView(View):
    template_name = 'admin_m/product_import_results.html'

    def post(self, request, *args, **kwargs):
        form = AdminImportFileForm(request.POST, request.FILES)
        import_summary = {'processed': 0, 'updated': 0, 'skipped': 0, 'failed_rows': [], 'created_values': 0}

        if form.is_valid():
            excel_file = request.FILES['file']
            logger.info(f"Processing uploaded Excel file: {excel_file.name}")

            try:
                workbook = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
                sheet = workbook.active
                headers_excel = [str(cell.value).strip() if cell.value is not None else '' for cell in sheet[1]]
                logger.debug(f"Excel Headers: {headers_excel}")

                # --- Pré-chargement et Mapping ---
                active_cats_by_slug = {fc.slug: fc for fc in FilterCategory.objects.filter(is_active=True)}
                header_map = {} # Pour SKU, Catalog ID, etc.
                filter_col_indices = {} # Pour {cat_slug: col_index}
                try:
                    header_map['sku'] = headers_excel.index('Артикул (SKU)')
                    header_map['catalog_id'] = headers_excel.index('ID Категории (для изменения)')
                except ValueError as e_col:
                    logger.error(f"Missing required column in Excel: {e_col}")
                    messages.error(request, f"Отсутствует обязательная колонка: {e_col}")
                    return redirect('admin_m:product_import')

                # --- Mapper les colonnes filtres ---
                header_slug_regex = re.compile(r'\(([^)]+)\)$') # Regex pour (...) à la fin
                for col_idx, header in enumerate(headers_excel):
                    # Vérifie si l'en-tête commence par "Фильтр:" ET s'il y a une correspondance regex
                    if header.startswith('Фильтр:'):
                        match = header_slug_regex.search(header.strip())
                        if match: # Traite seulement si la regex a trouvé quelque chose
                            cat_slug = match.group(1)
                            if cat_slug in active_cats_by_slug:
                                filter_col_indices[cat_slug] = col_idx
                                logger.debug(f"Mapped header '{header}' (Col {col_idx}) to filter cat '{cat_slug}'")
                            else:
                                logger.warning(f"Header '{header}' has inactive/unknown cat slug '{cat_slug}'. Column ignored.")
                        else:
                            # La regex n'a pas matché, log sans utiliser 'match' ou 'cat_slug'
                            logger.warning(f"Could not extract slug from header in parentheses: '{header}'. Expected '... (slug)'. Column ignored.")
                # ---------------------------------------------------------
                logger.info(f"Filter column mapping complete: {filter_col_indices}")

                # --- Itérer sur les Lignes ---
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    # ... (La logique interne de traitement de la ligne reste la même que celle
                    #      fournie dans la réponse précédente, qui était correcte pour
                    #      la mise à jour de catégorie et l'assignation des filtres M2M) ...

                    # --- DEBUT COPIE LOGIQUE LIGNE ---
                    import_summary['processed'] += 1
                    sku = str(row[header_map['sku']]).strip() if header_map['sku'] < len(row) and row[header_map['sku']] else None
                    new_cat_id_str = str(row[header_map['catalog_id']]).strip() if header_map['catalog_id'] < len(row) and row[header_map['catalog_id']] else None
                    log_prefix = f"Row {row_idx}: SKU {sku or 'N/A'} -"
                    if not sku: logger.warning(f"{log_prefix} Skipping: Missing SKU."); import_summary['skipped'] += 1; import_summary['failed_rows'].append({'row': row_idx, 'reason': 'Missing SKU'}); continue
                    try: product = Product.objects.select_related('catalog').prefetch_related('filters__category').get(sku=sku)
                    except Product.DoesNotExist: logger.warning(f"{log_prefix} Skipping: Not found"); import_summary['skipped'] += 1; import_summary['failed_rows'].append({'row': row_idx, 'sku': sku, 'reason': 'Not found'}); continue
                    except Exception as e_get: logger.error(f"{log_prefix} DB Error fetch: {e_get}"); import_summary['failed_rows'].append({'row': row_idx, 'sku': sku, 'reason': f'DB Error ({e_get})'}); continue

                    category_updated=False; filters_updated=False; save_fields=['updated_at']
                    target_catalog=None
                    if new_cat_id_str:
                        try: new_cat_id=int(new_cat_id_str)
                        except: pass
                        else:
                            if product.catalog_id != new_cat_id:
                                try: target_catalog = MenuCatalog.objects.get(pk=new_cat_id)
                                except MenuCatalog.DoesNotExist: logger.warning(f"{log_prefix} Target cat ID {new_cat_id} not found.")
                                else: product.catalog = target_catalog; category_updated = True; save_fields.append('catalog')

                    filter_pks_to_set=set(); filters_provided_in_row=False
                    for cat_slug, col_idx in filter_col_indices.items():
                        if col_idx < len(row) and row[col_idx] is not None:
                            values_str = str(row[col_idx]).strip(); filters_provided_in_row = True
                            if values_str:
                                filter_cat=active_cats_by_slug.get(cat_slug)
                                if filter_cat:
                                    values = [v.strip() for v in values_str.split('|') if v.strip()]
                                    for value_str in values:
                                        try: fv,cr=FilterValue.objects.get_or_create(category=filter_cat,value=value_str); filter_pks_to_set.add(fv.pk); import_summary['created_values'] += cr
                                        except Exception as e_fv: logger.error(f"{log_prefix} Error G/C FV '{value_str}': {e_fv}")

                    current_filter_pks = set(product.filters.values_list('pk', flat=True))
                    if filters_provided_in_row:
                        if current_filter_pks != filter_pks_to_set: filters_updated = True
                    # elif current_filter_pks: filters_updated = True # Pour effacer

                    something_changed = category_updated or filters_updated
                    if something_changed:
                        try:
                            if len(save_fields) > 1: product.save(update_fields=[f for f in save_fields if hasattr(product,f)])
                            if filters_updated:
                                if filters_provided_in_row: product.filters.set(list(filter_pks_to_set))
                                # elif current_filter_pks: product.filters.clear() # Pour effacer
                            import_summary['updated'] += 1; logger.info(f"{log_prefix} Updated OK (Cat:{category_updated}, Filt:{filters_updated})")
                        except Exception as e_save: logger.error(f"{log_prefix} Error saving: {e_save}"); import_summary['failed_rows'].append({'row':row_idx,'sku':sku,'reason':f'Save Error({e_save})'})
                    else: logger.debug(f"{log_prefix} No changes."); import_summary['skipped'] += 1
                    # ---- FIN Logique Copiée/Collée ----


                # --- Fin Boucle Lignes ---
                messages.success(request, f"Импорт завершен. Обработано: {import_summary['processed']}, Обновлено: {import_summary['updated']}, Пропущено: {import_summary['skipped']}, Ошибки: {len(import_summary['failed_rows'])}.")

            except Exception as e_process:
                logger.exception(f"Error processing Excel file {excel_file.name}: {e_process}")
                messages.error(request, f"Ошибка при обработке файла Excel: {e_process}")
                return redirect('admin_m:product_import')

        else: # Formulaire upload invalide
            logger.warning(f"Import file upload form invalid: {form.errors}")
            
            context = { 
                **get_base_admin_context(request, Product), 
                'title': 'Импорт Продуктов (Ошибки)', 
                'form': form 
            }
            return render(request, 'admin_m/product_import_setup.html', context)

        # --- Affichage Résultats ---
        context = { 
            **get_base_admin_context(request, Product), 
            'title': 'Результаты Импорта', 
            'summary': import_summary 
        }
        return render(request, self.template_name, context)