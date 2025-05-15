# admin_m/views.py
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import io
from datetime import datetime
from collections import defaultdict
import logging
import re

from django.shortcuts import render, redirect
from django.views import View
from django.http import HttpResponse, JsonResponse # JsonResponse si besoin
from django.contrib.admin.views.decorators import staff_member_required # Sécuriser l'accès
from django.utils.decorators import method_decorator
from django.urls import reverse
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django import forms
from django.contrib import messages
from django.db import transaction
from django.utils.text import slugify
from unidecode import unidecode
from django.contrib import admin # Importer admin pour accéder à site.site_header etc.

# Importer les modèles de l'application 'menu' (ou autre)
from menu.models import Product, MenuCatalog, FilterCategory, FilterValue
# Importer les constantes
from menu.admin import ID_UNCATEGORIZED_CATEGORY, ID_TYPE_CATEGORY_LIST , ID_TYPE_CATEGORY_LIST_2 # Ou depuis settings

logger = logging.getLogger(__name__)

# --- Formulaires pour les vues personnalisées ---
class AdminExportByCategoryForm(forms.Form):
    category = forms.ModelChoiceField(
        # queryset=MenuCatalog.objects.filter(is_hidden=False).exclude(id=ID_UNCATEGORIZED_CATEGORY).order_by('name'),
        queryset=MenuCatalog.objects.filter(type_menu_id__in=[ID_TYPE_CATEGORY_LIST, ID_TYPE_CATEGORY_LIST_2]).order_by('name'),
        label="Категория для экспорта",
        required=True,
        widget=forms.Select(attrs={'class': 'form-control'}) # Ajout classe pour style potentiel
    )
    # Ajoutez d'autres options si nécessaire (ex: format)

class AdminImportFileForm(forms.Form):
    file = forms.FileField(label="Файл Excel (.xlsx)", required=True,
                        widget=forms.FileInput(attrs={'accept': '.xlsx'}))
    # Option: Ajouter un choix pour le dry_run
    # dry_run = forms.BooleanField(label="Тестовый прогон (без сохранения)", required=False, initial=True)


# --- Vues ---

@method_decorator(staff_member_required, name='dispatch')
class ProductExportSetupView(View):
    form_class = AdminExportByCategoryForm
    template_name = 'admin_m/product_export_setup.html'

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        context = {
            # --- Contexte Admin Essentiel (Manuel) ---
            'title': 'Экспорт Продуктов по Категории',
            'site_header': admin.site.site_header, # Récupère depuis l'instance admin
            'site_title': admin.site.site_title,   # Récupère depuis l'instance admin
            'has_permission': True, # Car staff_member_required
            'opts': Product._meta, # Passe les meta-infos du modèle Product
            # -----------------------------------------
            'form': form,
            # 'app_list': admin.site.get_app_list(request), # Si vous avez besoin de la liste des apps dans base_site.html
        }
        return render(request, self.template_name, context)

@method_decorator(staff_member_required, name='dispatch')
class ProductExportDownloadView(View):
    """Génère et télécharge le fichier Excel basé sur la catégorie choisie."""
    resource_class = None # Pas besoin de django-import-export ici

    def get(self, request, *args, **kwargs):
        # Récupère la catégorie depuis les paramètres GET (envoyés par le formulaire précédent)
        category_id = request.GET.get('category')
        if not category_id:
            messages.error(request, "Категория не выбрана.")
            return redirect('admin_m:product_export_setup')

        try:
            category = MenuCatalog.objects.get(pk=category_id)
        except (MenuCatalog.DoesNotExist, ValueError):
            messages.error(request, "Выбранная категория не найдена.")
            return redirect('admin_m:product_export_setup')

        logger.info(f"Generating Excel export for category: {category.name}")

        # Récupère les produits
        products_to_export = Product.objects.filter(catalog=category).order_by('sku')
        products_to_export = products_to_export.prefetch_related('filters__category') # Précharge pour filtres

        if not products_to_export.exists():
            messages.warning(request, f"В категории '{category.name}' нет продуктов для экспорта.")
            return redirect('admin_m:product_export_setup')

        # --- Génération Excel avec openpyxl ---
        try:
            excel_buffer = io.BytesIO()
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = category.name[:30] # Limite longueur nom feuille

            # En-têtes (Colonnes nécessaires pour l'import)
            headers = ['Артикул (SKU)', 'Название', 'ID Категории (для изменения)', 'Ед. изм.']
            # Ajout dynamique des en-têtes de filtre
            filter_categories = FilterCategory.objects.filter(is_active=True).order_by('order', 'name')
            filter_slug_map = {} # Pour retrouver le slug dans la boucle produit
            for fc in filter_categories:
                column_name = f'Фильтр: {fc.name} ({fc.slug})'
                headers.append(column_name)
                filter_slug_map[fc.slug] = column_name

            sheet.append(headers)
            # ... (Application styles headers si besoin) ...

            # Écriture des données produit
            for product in products_to_export.iterator(chunk_size=1000):
                # Regroupe les filtres de CE produit par slug de catégorie
                product_filters = defaultdict(list)
                for fv in product.filters.filter(category__is_active=True): # Itère sur les filtres préchargés
                    product_filters[fv.category.slug].append(fv.value)

                row_data = [
                    product.sku,
                    product.title,
                    product.catalog_id, # ID actuel
                    product.ed_izm,
                ]
                # Ajoute les valeurs de filtre dans le bon ordre de colonne
                for fc in filter_categories: # Itère dans le même ordre que les headers
                    values = product_filters.get(fc.slug, [])
                    row_data.append("|".join(sorted(values))) # Ajoute les valeurs jointes

                sheet.append(row_data)

            workbook.save(excel_buffer)
            excel_buffer.seek(0)

            # --- Création Réponse HTTP ---
            response = HttpResponse(
                excel_buffer.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            timestamp = datetime.now().strftime('%Y%m%d_%H%M')
            filename = f"EXPORT_cat_{slugify(unidecode(category.name))}_{timestamp}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            logger.info(f"Excel export generated: {filename}")
            excel_buffer.close()
            return response

        except Exception as e_excel:
            logger.exception(f"Error generating Excel export for category {category.name}: {e_excel}")
            messages.error(request, "Произошла ошибка при создании файла Excel.")
            return redirect('admin_m:product_export_setup')

@method_decorator(staff_member_required, name='dispatch')
class ProductImportView(View):
    form_class = AdminImportFileForm
    template_name = 'admin_m/product_import_setup.html'

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        context = {
            # --- Contexte Admin Essentiel (Manuel) ---
            'title': 'Импорт Продуктов (Категории/Фильтры)',
            'site_header': admin.site.site_header,
            'site_title': admin.site.site_title,
            'has_permission': True,
            'opts': Product._meta,
            # -----------------------------------------
            'form': form,
            # 'app_list': admin.site.get_app_list(request), # Si vous avez besoin de la liste des apps dans base_site.html
        }
        return render(request, self.template_name, context)

@method_decorator(staff_member_required, name='dispatch')
@method_decorator(transaction.atomic, name='dispatch') # Enveloppe le traitement dans une transaction
class ProductImportProcessView(View):
    """Traite le fichier Excel uploadé."""
    template_name = 'admin_m/product_import_results.html' # Pour afficher les résultats

    def post(self, request, *args, **kwargs):
        form = AdminImportFileForm(request.POST, request.FILES)
        import_summary = {'updated': 0, 'created': 0, 'skipped': 0, 'failed_rows': [], 'processed': 0}

        if form.is_valid():
            excel_file = request.FILES['file']
            logger.info(f"Processing uploaded Excel file: {excel_file.name}")

            try:
                workbook = openpyxl.load_workbook(excel_file, read_only=True)
                sheet = workbook.active

                # Lire les en-têtes pour mapper les colonnes
                headers_excel = [str(cell.value).strip() if cell.value is not None else '' for cell in sheet[1]]
                logger.debug(f"Excel Headers: {headers_excel}")

                # Mapper les en-têtes attendus aux index de colonnes
                try:
                    sku_col = headers_excel.index('Артикул (SKU)')
                    cat_id_col = headers_excel.index('ID Категории (для изменения)')
                    # Ajoutez d'autres colonnes de base si vous les importez (ex: ed_izm)
                    # ed_izm_col = headers_excel.index('Ед. изм.')
                except ValueError as e_col:
                    logger.error(f"Missing required column in Excel: {e_col}")
                    messages.error(request, f"Отсутствует обязательная колонка в файле Excel: {e_col}")
                    return redirect('admin_m:product_import')

                # Mapper les colonnes de filtres dynamiques
                filter_col_map = {} # slug_cat -> index_col
                active_cats = {fc.slug: fc for fc in FilterCategory.objects.filter(is_active=True)}
                for col_idx, header in enumerate(headers_excel):
                    if header.startswith('Фильтр:'):
                        # Essayer d'extraire le slug de la catégorie depuis l'en-tête
                        match = re.search(r'\(slug=([a-zA-Z0-9_-]+)\)', header)
                        if match:
                            cat_slug = match.group(1)
                            if cat_slug in active_cats:
                                filter_col_map[cat_slug] = col_idx
                        else:
                            logger.warning(f"Could not extract category slug from filter header: '{header}'")


                # Itérer sur les lignes (commence à la ligne 2)
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    import_summary['processed'] += 1
                    sku = str(row[sku_col]).strip() if row[sku_col] else None
                    new_cat_id_str = str(row[cat_id_col]).strip() if row[cat_id_col] else None
                    # ed_izm_val = str(row[ed_izm_col]).strip() if len(row)>ed_izm_col and row[ed_izm_col] else None

                    if not sku:
                        logger.warning(f"Row {row_idx}: Skipping row due to missing SKU.")
                        import_summary['skipped'] += 1
                        import_summary['failed_rows'].append({'row': row_idx, 'reason': 'Missing SKU'})
                        continue

                    try:
                        product = Product.objects.get(sku=sku)
                        created = False
                    except Product.DoesNotExist:
                        # Optionnel : Créer le produit s'il n'existe pas ?
                        # Pour l'instant, on saute s'il n'existe pas car on se concentre sur l'assignation
                        logger.warning(f"Row {row_idx}: Product with SKU '{sku}' not found. Skipping.")
                        import_summary['skipped'] += 1
                        import_summary['failed_rows'].append({'row': row_idx, 'sku': sku, 'reason': 'Product not found'})
                        continue
                    except Exception as e_get:
                        logger.error(f"Row {row_idx}: Error fetching product SKU '{sku}': {e_get}")
                        import_summary['failed_rows'].append({'row': row_idx, 'sku': sku, 'reason': f'DB Error ({e_get})'})
                        continue

                    # --- Mise à jour Catégorie ---
                    category_updated = False
                    if new_cat_id_str:
                        try:
                            new_cat_id = int(new_cat_id_str)
                            if product.catalog_id != new_cat_id:
                                try:
                                    new_catalog = MenuCatalog.objects.get(pk=new_cat_id)
                                    product.catalog = new_catalog
                                    category_updated = True
                                    logger.debug(f"Row {row_idx}: SKU {sku} - Category will be updated to ID {new_cat_id}")
                                except MenuCatalog.DoesNotExist:
                                    logger.warning(f"Row {row_idx}: SKU {sku} - Target category ID {new_cat_id} not found. Category not updated.")
                                    import_summary['failed_rows'].append({'row': row_idx, 'sku': sku, 'reason': f'Target Category ID {new_cat_id} not found'})
                        except (ValueError, TypeError):
                            logger.warning(f"Row {row_idx}: SKU {sku} - Invalid category ID format: '{new_cat_id_str}'. Category not updated.")
                            import_summary['failed_rows'].append({'row': row_idx, 'sku': sku, 'reason': f'Invalid Category ID {new_cat_id_str}'})

                    # --- Mise à jour Autres Champs (Ex: ed_izm) ---
                    # if ed_izm_val is not None and product.ed_izm != ed_izm_val:
                    #     product.ed_izm = ed_izm_val
                    #     # Marquer pour sauvegarde même si catégorie/filtres ne changent pas
                    #     if not category_updated: save_needed = True # A définir plus haut

                    # --- Préparation des Filtres ---
                    filters_to_set_pks = set()
                    filters_found_in_row = False # Pour savoir si on doit potentiellement effacer
                    for cat_slug, col_idx in filter_col_map.items():
                        if len(row) > col_idx and row[col_idx] is not None:
                            values_str = str(row[col_idx]).strip()
                            if values_str:
                                filters_found_in_row = True # Au moins une colonne filtre a des données
                                filter_cat = active_cats.get(cat_slug)
                                if filter_cat:
                                    values = [v.strip() for v in values_str.split('|') if v.strip()]
                                    for value_str in values:
                                        try:
                                            fv, created = FilterValue.objects.get_or_create(category=filter_cat, value=value_str)
                                            filters_to_set_pks.add(fv.pk)
                                            if created: logger.info(f"Row {row_idx}: SKU {sku} - Created FV '{value_str}' for '{cat_slug}'")
                                        except Exception as e_fv: logger.error(f"Row {row_idx}: SKU {sku} - Error get/create FV '{value_str}': {e_fv}")
                                else:
                                    logger.warning(f"Row {row_idx}: SKU {sku} - Filter category slug '{cat_slug}' from header not found or inactive.")


                    # --- Comparaison et Assignation Filtres M2M ---
                    filters_updated = False
                    current_filter_pks = set(product.filters.values_list('pk', flat=True))
                    if filters_found_in_row: # Si des filtres étaient présents dans la ligne
                        if current_filter_pks != filters_to_set_pks:
                            product.filters.set(list(filters_to_set_pks))
                            filters_updated = True
                            logger.debug(f"Row {row_idx}: SKU {sku} - Filters updated.")
                    elif current_filter_pks: # Si pas de filtre dans la ligne mais il y en avait avant
                        product.filters.clear()
                        filters_updated = True
                        logger.debug(f"Row {row_idx}: SKU {sku} - Filters cleared.")

                    # --- Sauvegarde si Changement ---
                    if category_updated or filters_updated: # Ou si d'autres champs ont changé (save_needed)
                        try:
                            # La méthode save() du produit met à jour availability_changed_at si 'available' changeait
                            # Ici, on sauvegarde juste la catégorie et les filtres M2M ont déjà été sauvés par .set/.clear
                            if category_updated:
                                product.save(update_fields=['catalog', 'updated_at']) # Sauvegarde que la catégorie
                            logger.info(f"Row {row_idx}: SKU {sku} - Updated (Category: {category_updated}, Filters: {filters_updated})")
                            import_summary['updated'] += 1
                        except Exception as e_save:
                            logger.error(f"Row {row_idx}: SKU {sku} - Error saving product: {e_save}")
                            import_summary['failed_rows'].append({'row': row_idx, 'sku': sku, 'reason': f'Save Error ({e_save})'})
                    else:
                        logger.debug(f"Row {row_idx}: SKU {sku} - No changes detected.")
                        import_summary['skipped'] += 1


                # --- Fin Boucle Lignes ---
                messages.success(request, f"Импорт завершен. Обработано: {import_summary['processed']}, Обновлено: {import_summary['updated']}, Пропущено: {import_summary['skipped']}, Ошибки: {len(import_summary['failed_rows'])}.")

            except Exception as e_process:
                logger.exception(f"Error processing Excel file {excel_file.name}: {e_process}")
                messages.error(request, f"Ошибка при обработке файла Excel: {e_process}")
                return redirect('admin_m:product_import')

        else: # Formulaire d'upload invalide
            logger.warning(f"Import file upload form invalid: {form.errors}")
            messages.error(request, "Ошибка загрузки файла. Пожалуйста, выберите файл Excel (.xlsx).")
            return redirect('admin_m:product_import')

        # --- Affichage des Résultats ---
        context = {
            # --- Contexte Admin Essentiel (Manuel) ---
            'title': 'Результаты Импорта',
            'site_header': admin.site.site_header,
            'site_title': admin.site.site_title,
            'has_permission': True,
            'opts': Product._meta,
            # Ajouter app_label pour certains liens admin dans base_site
            'app_label': Product._meta.app_label,
        # -----------------------------------------
            'summary': import_summary,
        }
        return render(request, self.template_name, context)