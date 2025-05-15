# admin_m/views.py
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import io
from datetime import datetime
from collections import defaultdict
import logging
import re

from django.shortcuts import render, redirect
from django.views import View
from django.http import HttpResponse, JsonResponse # JsonResponse si besoin
from django.contrib.admin.views.decorators import staff_member_required # Sécuriser l'accès
from django.utils.decorators import method_decorator
from django.urls import reverse
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django import forms
from django.contrib import messages
from django.db import transaction, IntegrityError
from django.utils.text import slugify
from unidecode import unidecode
from django.contrib import admin # Importer admin pour accéder à site.site_header etc.
from collections import defaultdict

# Importer les modèles de l'application 'menu' (ou autre)
from menu.models import Product, MenuCatalog, FilterCategory, FilterValue
# Importer les constantes
from menu.admin import ID_UNCATEGORIZED_CATEGORY, ID_TYPE_CATEGORY_LIST , ID_TYPE_CATEGORY_LIST_2 # Ou depuis settings

logger = logging.getLogger(__name__)

# --- Formulaires pour les vues personnalisées ---
class AdminExportByCategoryForm(forms.Form):
    category = forms.ModelChoiceField(
        # queryset=MenuCatalog.objects.filter(is_hidden=False).exclude(id=ID_UNCATEGORIZED_CATEGORY).order_by('name'),
        queryset=MenuCatalog.objects.filter(type_menu_id__in=[ID_TYPE_CATEGORY_LIST, ID_TYPE_CATEGORY_LIST_2]).order_by('name'),
        label="Категория для экспорта",
        required=True,
        widget=forms.Select(attrs={'class': 'form-control'}) # Ajout classe pour style potentiel
    )
    # Ajoutez d'autres options si nécessaire (ex: format)

class AdminImportFileForm(forms.Form):
    file = forms.FileField(label="Файл Excel (.xlsx)", required=True,
                        widget=forms.FileInput(attrs={'accept': '.xlsx'}))
    # Option: Ajouter un choix pour le dry_run
    # dry_run = forms.BooleanField(label="Тестовый прогон (без сохранения)", required=False, initial=True)


# --- Vues ---

@method_decorator(staff_member_required, name='dispatch')
class ProductExportSetupView(View):
    form_class = AdminExportByCategoryForm
    template_name = 'admin_m/product_export_setup.html'

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        context = {
            # --- Contexte Admin Essentiel (Manuel) ---
            'title': 'Экспорт Продуктов по Категории',
            'site_header': admin.site.site_header, # Récupère depuis l'instance admin
            'site_title': admin.site.site_title,   # Récupère depuis l'instance admin
            'has_permission': True, # Car staff_member_required
            'opts': Product._meta, # Passe les meta-infos du modèle Product
            # -----------------------------------------
            'form': form,
            # 'app_list': admin.site.get_app_list(request), # Si vous avez besoin de la liste des apps dans base_site.html
        }
        return render(request, self.template_name, context)

@method_decorator(staff_member_required, name='dispatch')
class ProductExportDownloadView(View):
    """Génère et télécharge le fichier Excel basé sur la catégorie choisie."""
    resource_class = None # Pas besoin de django-import-export ici

    def get(self, request, *args, **kwargs):
        # Récupère la catégorie depuis les paramètres GET (envoyés par le formulaire précédent)
        category_id = request.GET.get('category')
        if not category_id:
            messages.error(request, "Категория не выбрана.")
            return redirect('admin_m:product_export_setup')

        try:
            category = MenuCatalog.objects.get(pk=category_id)
        except (MenuCatalog.DoesNotExist, ValueError):
            messages.error(request, "Выбранная категория не найдена.")
            return redirect('admin_m:product_export_setup')

        logger.info(f"Generating Excel export for category: {category.name}")

        # Récupère les produits
        products_to_export = Product.objects.filter(catalog=category).order_by('sku')
        products_to_export = products_to_export.prefetch_related('filters__category') # Précharge pour filtres

        if not products_to_export.exists():
            messages.warning(request, f"В категории '{category.name}' нет продуктов для экспорта.")
            return redirect('admin_m:product_export_setup')

        # --- Génération Excel avec openpyxl ---
        try:
            excel_buffer = io.BytesIO()
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = category.name[:30] # Limite longueur nom feuille

            # En-têtes (Colonnes nécessaires pour l'import)
            headers = ['Артикул (SKU)', 'Название', 'ID Категории (для изменения)', 'Ед. изм.']
            # Ajout dynamique des en-têtes de filtre
            filter_categories = FilterCategory.objects.filter(is_active=True).order_by('order', 'name')
            filter_slug_map = {} # Pour retrouver le slug dans la boucle produit
            for fc in filter_categories:
                column_name = f'Фильтр: {fc.name} ({fc.slug})'
                headers.append(column_name)
                filter_slug_map[fc.slug] = column_name

            sheet.append(headers)
            # ... (Application styles headers si besoin) ...

            # Écriture des données produit
            for product in products_to_export.iterator(chunk_size=1000):
                # Regroupe les filtres de CE produit par slug de catégorie
                product_filters = defaultdict(list)
                for fv in product.filters.filter(category__is_active=True): # Itère sur les filtres préchargés
                    product_filters[fv.category.slug].append(fv.value)

                row_data = [
                    product.sku,
                    product.title,
                    product.catalog_id, # ID actuel
                    product.ed_izm,
                ]
                # Ajoute les valeurs de filtre dans le bon ordre de colonne
                for fc in filter_categories: # Itère dans le même ordre que les headers
                    values = product_filters.get(fc.slug, [])
                    row_data.append("|".join(sorted(values))) # Ajoute les valeurs jointes

                sheet.append(row_data)

            workbook.save(excel_buffer)
            excel_buffer.seek(0)

            # --- Création Réponse HTTP ---
            response = HttpResponse(
                excel_buffer.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            timestamp = datetime.now().strftime('%Y%m%d_%H%M')
            filename = f"EXPORT_cat_{slugify(unidecode(category.name))}_{timestamp}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            logger.info(f"Excel export generated: {filename}")
            excel_buffer.close()
            return response

        except Exception as e_excel:
            logger.exception(f"Error generating Excel export for category {category.name}: {e_excel}")
            messages.error(request, "Произошла ошибка при создании файла Excel.")
            return redirect('admin_m:product_export_setup')

@method_decorator(staff_member_required, name='dispatch')
class ProductImportView(View):
    form_class = AdminImportFileForm
    template_name = 'admin_m/product_import_setup.html'

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        context = {
            # --- Contexte Admin Essentiel (Manuel) ---
            'title': 'Импорт Продуктов (Категории/Фильтры)',
            'site_header': admin.site.site_header,
            'site_title': admin.site.site_title,
            'has_permission': True,
            'opts': Product._meta,
            # -----------------------------------------
            'form': form,
            # 'app_list': admin.site.get_app_list(request), # Si vous avez besoin de la liste des apps dans base_site.html
        }
        return render(request, self.template_name, context)

# # --- Vue de Traitement de l'Import ---
# @method_decorator(staff_member_required, name='dispatch')
# @method_decorator(transaction.atomic, name='dispatch') # Enveloppe le POST dans une transaction
# class ProductImportProcessView(View):
#     template_name = 'admin_m/product_import_results.html' # Pour afficher les résultats

#     def post(self, request, *args, **kwargs):
#         form = AdminImportFileForm(request.POST, request.FILES)
#         # Initialise le résumé
#         import_summary = {'processed': 0, 'updated': 0, 'skipped': 0, 'failed_rows': [], 'created_values': 0}

#         if form.is_valid():
#             excel_file = request.FILES['file']
#             logger.info(f"Processing uploaded Excel file: {excel_file.name}")

#             try:
#                 workbook = openpyxl.load_workbook(excel_file, read_only=True, data_only=True) # data_only pour lire les valeurs
#                 sheet = workbook.active

#                 # --- Lecture et Mapping des En-têtes ---
#                 headers_excel = [str(cell.value).strip() if cell.value is not None else '' for cell in sheet[1]]
#                 logger.debug(f"Excel Headers: {headers_excel}")

#                 header_map = {} # Stocke {'nom_logique': index_colonne}
#                 filter_col_indices = {} # Stocke {category_slug: index_colonne}
#                 active_cats_by_slug = {fc.slug: fc for fc in FilterCategory.objects.filter(is_active=True)} # Précharge

#                 # Vérifie les colonnes obligatoires et optionnelles
#                 required_cols = {'Артикул (SKU)': 'sku', 'ID Категории (для изменения)': 'catalog_id'}
#                 optional_cols = {'Ед. изм.': 'ed_izm'} # Ajoutez d'autres si importés

#                 for req_col, field_key in required_cols.items():
#                     try: header_map[field_key] = headers_excel.index(req_col)
#                     except ValueError:
#                         logger.error(f"Missing required column in Excel: '{req_col}'")
#                         messages.error(request, f"Отсутствует обязательная колонка: '{req_col}'")
#                         return redirect('admin_m:product_import')

#                 for opt_col, field_key in optional_cols.items():
#                     try: header_map[field_key] = headers_excel.index(opt_col)
#                     except ValueError: logger.warning(f"Optional column '{opt_col}' not found.")

#                 # Mapper les colonnes filtres dynamiques
#                 for col_idx, header in enumerate(headers_excel):
#                     if header.startswith('Фильтр:'):
#                          match = re.search(r'\(slug=([a-zA-Z0-9_-]+)\)', header)
#                          if match:
#                              cat_slug = match.group(1)
#                              if cat_slug in active_cats_by_slug:
#                                  filter_col_indices[cat_slug] = col_idx
#                                  logger.debug(f"Mapped header '{header}' (Col {col_idx}) to filter cat '{cat_slug}'")
#                              else: logger.warning(f"Header '{header}' has inactive/unknown cat slug '{cat_slug}'. Ignoring.")
#                          else: logger.warning(f"Could not extract slug from header: '{header}'. Ignoring.")
#                 # --- Fin Mapping En-têtes ---


#                 # --- Itérer sur les Lignes de Données ---
#                 for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
#                     import_summary['processed'] += 1
#                     # Utilise get() avec index pour éviter IndexError si ligne courte
#                     sku = str(row[header_map['sku']]).strip() if header_map['sku'] < len(row) and row[header_map['sku']] else None
#                     new_cat_id_str = str(row[header_map['catalog_id']]).strip() if header_map['catalog_id'] < len(row) and row[header_map['catalog_id']] else None
#                     # ed_izm_val = str(row[header_map['ed_izm']]).strip() if 'ed_izm' in header_map and header_map['ed_izm'] < len(row) and row[header_map['ed_izm']] else None

#                     log_prefix = f"Row {row_idx}: SKU {sku or 'N/A'} -" # Pour logs

#                     if not sku:
#                         logger.warning(f"{log_prefix} Skipping due to missing SKU.")
#                         import_summary['skipped'] += 1; import_summary['failed_rows'].append({'row': row_idx, 'reason': 'Missing SKU'}); continue

#                     # --- Récupération Produit Existant ---
#                     try:
#                         product = Product.objects.select_related('catalog').prefetch_related('filters__category').get(sku=sku)
#                         created = False
#                     except Product.DoesNotExist:
#                         logger.warning(f"{log_prefix} Product not found. Skipping row (creation not implemented here).")
#                         import_summary['skipped'] += 1; import_summary['failed_rows'].append({'row': row_idx, 'sku': sku, 'reason': 'Product not found'}); continue
#                     except Exception as e_get:
#                          logger.error(f"{log_prefix} DB Error fetching product: {e_get}")
#                          import_summary['failed_rows'].append({'row': row_idx, 'sku': sku, 'reason': f'DB Error ({e_get})'}); continue

#                     # Flags pour suivre les changements sur CET objet
#                     category_changed = False
#                     filters_changed = False
#                     other_fields_changed = False # Si vous importez d'autres champs

#                     # --- 1. Mise à Jour Catégorie ---
#                     target_catalog = None
#                     if new_cat_id_str:
#                         try:
#                             new_cat_id = int(new_cat_id_str)
#                             if product.catalog_id != new_cat_id:
#                                 try:
#                                     target_catalog = MenuCatalog.objects.get(pk=new_cat_id) # Valide que la catégorie existe
#                                     product.catalog = target_catalog
#                                     category_changed = True
#                                     logger.debug(f"{log_prefix} Category marked for update to ID {new_cat_id}")
#                                 except MenuCatalog.DoesNotExist:
#                                      logger.warning(f"{log_prefix} Target category ID {new_cat_id} not found.")
#                                      import_summary['failed_rows'].append({'row': row_idx, 'sku': sku, 'reason': f'Target Category ID {new_cat_id} not found'})
#                         except (ValueError, TypeError):
#                              logger.warning(f"{log_prefix} Invalid category ID format: '{new_cat_id_str}'.")
#                              import_summary['failed_rows'].append({'row': row_idx, 'sku': sku, 'reason': f'Invalid Category ID {new_cat_id_str}'})

#                     # --- 2. Mise à Jour Autres Champs (Ex: ed_izm) ---
#                     # if 'ed_izm' in header_map and ed_izm_val is not None:
#                     #     if product.ed_izm != ed_izm_val:
#                     #         product.ed_izm = ed_izm_val
#                     #         other_fields_changed = True
#                     #         logger.debug(f"{log_prefix} ed_izm marked for update.")

#                     # --- 3. Traitement et Comparaison des Filtres ---
#                     filter_pks_to_set = set()
#                     filters_provided_in_row = False # Réinitialisé pour chaque ligne
#                     for cat_slug, col_idx in filter_col_indices.items():
#                         if col_idx < len(row) and row[col_idx] is not None:
#                             values_str = str(row[col_idx]).strip()
#                             filters_provided_in_row = True # La colonne existe et n'est pas None
#                             if values_str: # Non vide après strip
#                                 filter_cat = active_cats_by_slug.get(cat_slug)
#                                 if filter_cat:
#                                     values = [v.strip() for v in values_str.split('|') if v.strip()]
#                                     for value_str in values:
#                                         try:
#                                             fv, created = FilterValue.objects.get_or_create(category=filter_cat, value=value_str)
#                                             filter_pks_to_set.add(fv.pk)
#                                             if created: import_summary['created_values'] += 1
#                                         except Exception as e_fv: logger.error(f"{log_prefix} Error G/C FV '{value_str}': {e_fv}")

#                     # Compare avec les filtres actuels
#                     current_filter_pks = set(product.filters.values_list('pk', flat=True))
#                     if filters_provided_in_row: # Si des données filtres étaient dans la ligne
#                          if current_filter_pks != filter_pks_to_set:
#                               filters_changed = True
#                               logger.debug(f"{log_prefix} Filters marked for update. Current: {current_filter_pks}, New: {filter_pks_to_set}")
#                     # Si vous voulez effacer les filtres quand aucune colonne filtre n'est fournie
#                     # elif current_filter_pks:
#                     #      filters_changed = True
#                     #      logger.debug(f"{log_prefix} Filters marked for clearing.")

#                     # --- 4. Sauvegarde Conditionnelle et Assignation M2M ---
#                     if category_changed or other_fields_changed or filters_changed:
#                         try:
#                             save_update_fields = ['updated_at'] # updated_at est géré par auto_now
#                             if category_changed: save_update_fields.append('catalog')
#                             # if other_fields_changed: save_update_fields.append('ed_izm'); # ...
#                             # Sauvegarde les champs simples si changés
#                             if len(save_update_fields) > 1:
#                                  logger.info(f"{log_prefix} Saving main fields: {save_update_fields}")
#                                  product.save(update_fields=save_update_fields)

#                             # Gère les filtres M2M SI ils ont changé
#                             if filters_changed:
#                                  if filters_provided_in_row: # Applique le nouvel ensemble
#                                      logger.info(f"{log_prefix} Setting M2M filters to PKs: {filter_pks_to_set}")
#                                      product.filters.set(list(filter_pks_to_set))
#                                  # Décommentez si vous voulez effacer quand aucune colonne n'est fournie
#                                  # elif current_filter_pks:
#                                  #      logger.info(f"{log_prefix} Clearing existing M2M filters.")
#                                  #      product.filters.clear()

#                             import_summary['updated'] += 1
#                             import_summary['updated_with_changes'] += 1
#                             logger.info(f"{log_prefix} Updated successfully (Category: {category_changed}, Filters: {filters_changed})")

#                         except IntegrityError as e_int: # Ex: Contrainte unique violée ailleurs
#                             logger.error(f"{log_prefix} IntegrityError saving product/M2M: {e_int}")
#                             import_summary['failed_rows'].append({'row': row_idx, 'sku': sku, 'reason': f'Integrity Error ({e_int})'})
#                         except Exception as e_save_m2m:
#                             logger.error(f"{log_prefix} Error saving product or M2M filters: {e_save_m2m}")
#                             import_summary['failed_rows'].append({'row': row_idx, 'sku': sku, 'reason': f'Save/M2M Error ({e_save_m2m})'})
#                             # La transaction atomique devrait annuler les changements partiels pour cette ligne
#                     else:
#                         logger.debug(f"{log_prefix} No changes detected. Skipping save.")
#                         import_summary['skipped'] += 1

#                 # --- Fin Boucle Lignes ---
#                 messages.success(request, f"Импорт завершен. Обработано: {import_summary['processed']}, Обновлено с изменениями: {import_summary['updated_with_changes']}, Пропущено (без изменений): {import_summary['skipped']}, Ошибки: {len(import_summary['failed_rows'])}.")

#             except Exception as e_process:
#                 logger.exception(f"Error processing Excel file {excel_file.name}: {e_process}")
#                 messages.error(request, f"Ошибка при обработке файла Excel: {e_process}")
#                 # Ne redirige pas pour pouvoir afficher le résumé partiel si besoin
#                 # return redirect('admin_m:product_import')

#         else: # Formulaire upload invalide
#             logger.warning(f"Import file upload form invalid: {form.errors}")
#             # Important : Ré-afficher le formulaire avec les erreurs
#             context = { 
#                 **admin.site.each_context(request), 
#                 'title': 'Импорт Продуктов (Ошибки)', 
#                 'form': form, 
#                 'opts': Product._meta, 
#                 'app_label': Product._meta.app_label
#                 }
#             return render(request, 'admin_m/product_import_setup.html', context)

#         # --- Affichage des Résultats ---
#         context = {
#             # --- Contexte Admin Essentiel (Manuel) ---
#             'title': 'Результаты Импорта',
#             'site_header': admin.site.site_header,
#             'site_title': admin.site.site_title,
#             'has_permission': True,
#             'opts': Product._meta,
#             # Ajouter app_label pour certains liens admin dans base_site
#             'app_label': Product._meta.app_label,
#         # -----------------------------------------
#             'summary': import_summary,
#         }
#         return render(request, self.template_name, context)


@method_decorator(staff_member_required, name='dispatch')
@method_decorator(transaction.atomic, name='dispatch')
class ProductImportProcessView(View):
    template_name = 'admin_m/product_import_results.html'

    def post(self, request, *args, **kwargs):
        form = AdminImportFileForm(request.POST, request.FILES)
        # --- Initialisation du résumé SANS 'updated_with_changes' ---
        import_summary = {'processed': 0, 'updated': 0, 'skipped': 0, 'failed_rows': [], 'created_values': 0}

        if form.is_valid():
            excel_file = request.FILES['file']
            logger.info(f"Processing uploaded Excel file: {excel_file.name}")

            try:
                workbook = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
                sheet = workbook.active
                headers_excel = [str(cell.value).strip() if cell.value is not None else '' for cell in sheet[1]]
                logger.debug(f"Excel Headers: {headers_excel}")

                # --- Pré-chargement et Mapping ---
                active_cats_by_slug = {fc.slug: fc for fc in FilterCategory.objects.filter(is_active=True)}
                header_map = {}
                filter_col_indices = {} # {cat_slug: col_index}
                try:
                    header_map['sku'] = headers_excel.index('Артикул (SKU)')
                    header_map['catalog_id'] = headers_excel.index('ID Категории (для изменения)')
                except ValueError as e_col:
                    logger.error(f"Missing required column in Excel: {e_col}")
                    messages.error(request, f"Отсутствует обязательная колонка: {e_col}")
                    return redirect('admin_m:product_import')

                # --- Mapper les colonnes filtres (REGEX CORRIGÉE) ---
                # Regex pour extraire le contenu entre les dernières parenthèses
                header_slug_regex = re.compile(r'\(([^)]+)\)$') # Cherche (...) à la fin
                for col_idx, header in enumerate(headers_excel):
                    if header.startswith('Фильтр:'):
                        match = header_slug_regex.search(header.strip())
                        if match:
                            # Le groupe 1 contient le slug (ex: 'proizvoditel')
                            cat_slug = match.group(1)
                            if cat_slug in active_cats_by_slug:
                                filter_col_indices[cat_slug] = col_idx
                                logger.debug(f"Mapped header '{header}' (Col {col_idx}) to filter cat '{cat_slug}'")
                            else:
                                logger.warning(f"Header '{header}' has inactive/unknown cat slug '{cat_slug}'. Column ignored.")
                        else:
                            # Log si le format (...) n'est pas trouvé à la fin
                            logger.warning(f"Could not extract slug from header in parentheses: '{header}'. Column ignored.")
                # ---------------------------------------------------------

                # --- Itérer sur les Lignes ---
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    import_summary['processed'] += 1
                    # ... (récupération sku, new_cat_id_str) ...
                    sku = str(row[header_map['sku']]).strip() if header_map['sku'] < len(row) and row[header_map['sku']] else None
                    new_cat_id_str = str(row[header_map['catalog_id']]).strip() if header_map['catalog_id'] < len(row) and row[header_map['catalog_id']] else None

                    if not sku: # ... (gestion SKU manquant) ...
                        continue

                    # --- Récupération Produit Existant ---
                    try: product = Product.objects.select_related('catalog').prefetch_related('filters__category').get(sku=sku)
                    except Product.DoesNotExist: # ... (gestion produit non trouvé) ...
                        continue
                    except Exception as e_get: # ... (gestion erreur DB) ...
                        continue

                    category_updated = False; filters_updated = False; other_fields_changed = False # Flags locaux
                    save_fields = ['updated_at'] # Commence avec updated_at

                    # --- 1. Mise à Jour Catégorie ---
                    # ... (Logique inchangée pour mettre à jour product.catalog et category_updated) ...
                    target_catalog = None
                    if new_cat_id_str:
                        try: new_cat_id = int(new_cat_id_str)
                        except (ValueError, TypeError): pass
                        else:
                            if product.catalog_id != new_cat_id:
                                try: target_catalog = MenuCatalog.objects.get(pk=new_cat_id)
                                except MenuCatalog.DoesNotExist: pass
                                else: product.catalog = target_catalog; category_updated = True; save_fields.append('catalog')


                    # --- 2. Traitement et Comparaison des Filtres ---
                    filter_pks_to_set = set()
                    filters_provided_in_row = False
                    for cat_slug, col_idx in filter_col_indices.items(): # Utilise la map créée
                        if col_idx < len(row) and row[col_idx] is not None:
                            values_str = str(row[col_idx]).strip()
                            filters_provided_in_row = True
                            if values_str:
                                filter_cat = active_cats_by_slug.get(cat_slug)
                                if filter_cat:
                                    values = [v.strip() for v in values_str.split('|') if v.strip()]
                                    for value_str in values:
                                        try:
                                            fv, created = FilterValue.objects.get_or_create(category=filter_cat, value=value_str)
                                            filter_pks_to_set.add(fv.pk)
                                            if created: import_summary['created_values'] += 1
                                        except Exception as e_fv: logger.error(f"Row {row_idx} SKU {sku}: Error G/C FV '{value_str}': {e_fv}")

                    current_filter_pks = set(product.filters.values_list('pk', flat=True))
                    if filters_provided_in_row:
                        if current_filter_pks != filter_pks_to_set:
                            filters_updated = True
                    # elif current_filter_pks: # Décommentez pour effacer si pas de colonnes fournies
                    #      filters_updated = True

                    # --- 3. Sauvegarde Conditionnelle ---
                    something_changed = category_updated or filters_updated # ou other_fields_changed
                    if something_changed:
                        try:
                            # Sauvegarde champs simples SI changés
                            if len(save_fields) > 1:
                                product.save(update_fields=[f for f in save_fields if hasattr(product, f)]) # Assure que le champ existe
                            # Applique changements M2M SI changés
                            if filters_updated:
                                if filters_provided_in_row:
                                    product.filters.set(list(filter_pks_to_set))
                                # elif current_filter_pks: # Logique d'effacement
                                #     product.filters.clear()
                            import_summary['updated'] += 1
                            logger.info(f"Row {row_idx}: SKU {sku} - Updated (Category: {category_updated}, Filters: {filters_updated})")
                        except Exception as e_save_m2m:
                            logger.error(f"Row {row_idx}: SKU {sku} - Error saving/M2M: {e_save_m2m}")
                            import_summary['failed_rows'].append({'row': row_idx, 'sku': sku, 'reason': f'Save/M2M Error ({e_save_m2m})'})
                    else:
                        logger.debug(f"Row {row_idx}: SKU {sku} - No changes detected.")
                        import_summary['skipped'] += 1


                # --- Fin Boucle Lignes ---
                # --- MODIFICATION : Message de succès SANS 'updated_with_changes' ---
                messages.success(request, f"Импорт завершен. Обработано: {import_summary['processed']}, Обновлено: {import_summary['updated']}, Пропущено: {import_summary['skipped']}, Ошибки: {len(import_summary['failed_rows'])}.")

            except Exception as e_process:
                logger.exception(f"Error processing Excel file {excel_file.name}: {e_process}")
                messages.error(request, f"Ошибка при обработке файла Excel: {e_process}")
                return redirect('admin_m:product_import')

        else: # Formulaire upload invalide
            logger.warning(f"Import file upload form invalid: {form.errors}")
            context = { **admin.site.each_context(request), 'title': 'Импорт Продуктов (Ошибки)', 'form': form, 'opts': Product._meta, 'app_label': Product._meta.app_label}
            return render(request, 'admin_m/product_import_setup.html', context)

        # --- Affichage Résultats ---
        context = { **admin.site.each_context(request), 'title': 'Результаты Импорта', 'summary': import_summary, 'opts': Product._meta, 'app_label': Product._meta.app_label }
        return render(request, self.template_name, context)